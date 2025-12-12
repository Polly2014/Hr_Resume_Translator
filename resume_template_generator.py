#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简历模板生成器
根据解析后的简历JSON数据，填充Excel模板生成标准化简历文档
"""

import json
import copy
import os
import re
from pathlib import Path
from typing import Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter


class ResumeTemplateGenerator:
    """简历模板生成器"""
    
    # 高亮样式 - 用于null值
    HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    HIGHLIGHT_FONT = Font(color="FF0000", bold=True)
    
    # 模板中的关键行号（基础位置，会随插入行动态调整）
    class RowPositions:
        NAME_SUPPLIER = 3           # 姓名、供应商
        PERSONAL_INFO_START = 5     # 个人信息开始
        UNDERGRADUATE_START = 14    # 本科学历开始
        GRADUATE_START = 22         # 研究生学历开始（原始位置）
        WORK_EXP_HEADER = 31        # 工作经历表头
        WORK_EXP_DATA = 33          # 工作经历数据开始
        PROJECT_EXP_HEADER = 36     # 项目经历表头
        PROJECT_EXP_DATA = 38       # 项目经历数据开始
        TECH_SKILLS_START = 41      # 技术特长开始
    
    def __init__(self, template_path: str):
        """
        初始化生成器
        
        Args:
            template_path: Excel模板文件路径
        """
        self.template_path = Path(template_path)
        self.wb = load_workbook(template_path)
        self.ws = self.wb.active
        self.row_offset = 0  # 跟踪因插入行导致的偏移
    
    @staticmethod
    def _format_date(date_str: Optional[str]) -> Optional[str]:
        """
        格式化日期为 YYYY-MM-DD 格式
        
        支持的输入格式：
        - YYYY-MM-DD → 原样返回
        - YYYY-MM → 补全为 YYYY-MM-01
        - YYYY → 补全为 YYYY-01-01
        - YYYY年MM月DD日 → 转换为 YYYY-MM-DD
        - YYYY年MM月 → 转换为 YYYY-MM-01
        - None/空 → 返回 None
        
        Args:
            date_str: 原始日期字符串
            
        Returns:
            格式化后的日期字符串，或 None
        """
        if not date_str or date_str.strip() == "":
            return None
        
        date_str = date_str.strip()
        
        # 已经是 YYYY-MM-DD 格式
        if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return date_str
        
        # YYYY-MM 格式，补全日期为 01
        if re.match(r'^\d{4}-\d{2}$', date_str):
            return f"{date_str}-01"
        
        # YYYY-M 格式（单位数月份），补全为 YYYY-0M-01
        if re.match(r'^\d{4}-\d{1}$', date_str):
            parts = date_str.split('-')
            return f"{parts[0]}-0{parts[1]}-01"
        
        # YYYY 格式，补全为 YYYY-01-01
        if re.match(r'^\d{4}$', date_str):
            return f"{date_str}-01-01"
        
        # 中文格式：YYYY年MM月DD日
        match = re.match(r'^(\d{4})年(\d{1,2})月(\d{1,2})日?$', date_str)
        if match:
            year, month, day = match.groups()
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # 中文格式：YYYY年MM月
        match = re.match(r'^(\d{4})年(\d{1,2})月$', date_str)
        if match:
            year, month = match.groups()
            return f"{year}-{month.zfill(2)}-01"
        
        # 其他格式，原样返回
        return date_str
        
    def _get_actual_row(self, base_row: int) -> int:
        """获取考虑偏移后的实际行号"""
        return base_row + self.row_offset
    
    def _fill_cell(self, row: int, col: int, value: Any, apply_offset: bool = True, is_date: bool = False) -> None:
        """
        填充单元格，null值会高亮显示
        
        Args:
            row: 行号
            col: 列号
            value: 要填充的值
            apply_offset: 是否应用行偏移
            is_date: 是否为日期字段，如果是则自动格式化
        """
        actual_row = self._get_actual_row(row) if apply_offset else row
        cell = self.ws.cell(row=actual_row, column=col)
        
        # 如果是日期字段，先格式化
        if is_date and value:
            value = self._format_date(value)
        
        if value is None or value == "" or value == []:
            cell.value = "【待补充】"
            cell.fill = self.HIGHLIGHT_FILL
            cell.font = self.HIGHLIGHT_FONT
        elif isinstance(value, list):
            cell.value = "、".join(str(v) for v in value) if value else "【待补充】"
            if not value:
                cell.fill = self.HIGHLIGHT_FILL
                cell.font = self.HIGHLIGHT_FONT
        elif isinstance(value, bool):
            cell.value = "是" if value else "否"
        else:
            cell.value = str(value)
    
    def _copy_row_style(self, source_row: int, target_row: int, num_cols: int = 15) -> None:
        """复制行样式"""
        for col in range(1, num_cols + 1):
            source_cell = self.ws.cell(row=source_row, column=col)
            target_cell = self.ws.cell(row=target_row, column=col)
            
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)
    
    def _insert_rows_with_style(self, start_row: int, num_rows: int, template_row: int) -> None:
        """插入行并复制样式"""
        if num_rows <= 0:
            return
            
        self.ws.insert_rows(start_row, num_rows)
        
        # 复制样式到新插入的行
        for i in range(num_rows):
            self._copy_row_style(template_row + num_rows, start_row + i)
        
        self.row_offset += num_rows
    
    def fill_basic_info(self, data: dict) -> None:
        """填充基本信息（姓名、供应商）"""
        basic_info = data.get("basic_info", {})
        
        # 姓名 - B3
        self._fill_cell(3, 2, basic_info.get("name"))
        # 供应商 - D3
        self._fill_cell(3, 4, basic_info.get("supplier"))
    
    def fill_personal_info(self, data: dict) -> None:
        """填充个人信息"""
        personal = data.get("personal_info", {})
        
        # 身份证号 - B5 (实际是第5行的第2列)
        self._fill_cell(5, 2, personal.get("id_card_number"))
        # 出生日期 - B7
        self._fill_cell(7, 2, personal.get("birth_date"), is_date=True)
        # 电话 - B8
        self._fill_cell(8, 2, personal.get("phone"))
        # 首次参加工作时间 - B9
        self._fill_cell(9, 2, personal.get("first_work_date"), is_date=True)
        # 首次参加IT领域工作时间 - B10
        self._fill_cell(10, 2, personal.get("first_it_work_date"), is_date=True)
        # 最高学历 - B11
        self._fill_cell(11, 2, personal.get("highest_education"))
        # 合同定级 - B12
        self._fill_cell(12, 2, personal.get("contract_level"))
    
    def fill_education(self, data: dict) -> None:
        """填充学历信息"""
        education_list = data.get("education", [])
        
        # 按学历类型分类
        undergraduate = None  # 本科
        master = None         # 硕士
        doctoral = None       # 博士
        
        for edu in education_list:
            degree_type = edu.get("degree_type", "")
            if "本科" in degree_type:
                undergraduate = edu
            elif "硕士" in degree_type:
                master = edu
            elif "博士" in degree_type:
                doctoral = edu
        
        # 填充本科学历 (行14-20)
        if undergraduate:
            self._fill_undergraduate(undergraduate)
        else:
            self._fill_undergraduate({})
        
        # 填充研究生学历 (行22-29)
        # 如果有硕士，填硕士；如果只有博士没有硕士，研究生区域填博士
        if master:
            self._fill_graduate(master, "硕士")
        elif doctoral and not master:
            # 只有博士没有硕士时，研究生区域填博士
            self._fill_graduate(doctoral, "博士")
            doctoral = None  # 已填充，不需要再创建博士区块
        
        # 如果同时有硕士和博士，需要插入博士学历区块
        if doctoral and master:
            self._insert_doctoral_section(doctoral)
    
    def _fill_undergraduate(self, edu: dict) -> None:
        """填充本科学历"""
        # 入学时间 - B14, 毕业院校 - D14
        self._fill_cell(14, 2, edu.get("enrollment_date"), is_date=True)
        self._fill_cell(14, 4, edu.get("university"))
        # 毕业时间 - B15, 专业 - D15
        self._fill_cell(15, 2, edu.get("graduation_date"), is_date=True)
        self._fill_cell(15, 4, edu.get("major"))
        # 毕业证编号 - B16
        self._fill_cell(16, 2, edu.get("diploma_number"))
        # 毕业证学信网验证码 - B17
        self._fill_cell(17, 2, edu.get("diploma_verification_code"))
        # 学位证编号 - B19
        self._fill_cell(19, 2, edu.get("degree_number"))
        # 学位证学信网验证码 - B20
        self._fill_cell(20, 2, edu.get("degree_verification_code"))
    
    def _fill_graduate(self, edu: dict, degree_label: str = "硕士") -> None:
        """填充研究生学历（硕士或博士）"""
        # 研究生区域从第22行开始
        base_row = 22
        
        # 更新标题（如果是博士）
        if degree_label == "博士":
            self.ws.cell(row=self._get_actual_row(base_row), column=1).value = "博士学历"
        
        # 入学时间 - B23, 毕业院校 - D23
        self._fill_cell(base_row + 1, 2, edu.get("enrollment_date"), is_date=True)
        self._fill_cell(base_row + 1, 4, edu.get("university"))
        # 毕业时间 - B24, 专业 - D24
        self._fill_cell(base_row + 2, 2, edu.get("graduation_date"), is_date=True)
        self._fill_cell(base_row + 2, 4, edu.get("major"))
        # 毕业证编号 - B25
        self._fill_cell(base_row + 3, 2, edu.get("diploma_number"))
        # 毕业证学信网验证码 - B26
        self._fill_cell(base_row + 4, 2, edu.get("diploma_verification_code"))
        # 学位证编号 - B28
        self._fill_cell(base_row + 6, 2, edu.get("degree_number"))
        # 学位证学信网验证码 - B29
        self._fill_cell(base_row + 7, 2, edu.get("degree_verification_code"))
    
    def _insert_doctoral_section(self, edu: dict) -> None:
        """在研究生学历下方插入博士学历区块"""
        # 博士学历区块结构（仿照研究生，共9行）：
        # Row 1: 博士学历（标题，合并单元格）
        # Row 2: 入学时间 | YYYY-MM-DD | 毕业院校 | xxxx
        # Row 3: 毕业时间 | YYYY-MM-DD | 专业 | xxxx
        # Row 4: 毕业证编号 | 18位数字
        # Row 5: 毕业证学信网在线验证码 | 16位字母+数字
        # Row 6: 空行
        # Row 7: 学位证编号 | 16位数字
        # Row 8: 学位证学信网在线验证码 | 16位字母+数字
        # Row 9: 空行
        
        # 插入位置：研究生学历后面（原始Row 30之后，即Row 31之前）
        # 但工作经历从Row 31开始，所以我们在Row 31插入9行
        insert_row = self._get_actual_row(31)
        num_rows = 9
        
        # 记录工作经历区域将被移动到的新位置
        work_exp_new_start = insert_row + num_rows  # 工作经历标题新位置
        work_exp_header_row = work_exp_new_start + 1  # 工作经历表头新位置
        
        # 插入空行
        self.ws.insert_rows(insert_row, num_rows)
        self.row_offset += num_rows
        
        # 修复工作经历区域被错误合并的单元格
        # insert_rows 可能导致错误的合并，需要清理
        merged_to_fix = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # 工作经历表头行（Row 32+9=41）如果被整行合并，需要修复
            if merged_range.min_row == work_exp_header_row:
                if merged_range.min_col == 1 and merged_range.max_col >= 3:
                    merged_to_fix.append(merged_range)
        
        for mr in merged_to_fix:
            try:
                self.ws.unmerge_cells(str(mr))
            except:
                pass
        
        # 重新正确合并工作经历表头的 E:F 列
        self.ws.merge_cells(
            start_row=work_exp_header_row, start_column=5,
            end_row=work_exp_header_row, end_column=6
        )
        
        # 恢复工作经历表头的值（insert_rows 可能导致值丢失）
        self.ws.cell(row=work_exp_header_row, column=1).value = "工作开始日期（年月日）"
        self.ws.cell(row=work_exp_header_row, column=2).value = "工作结束日期（年月日）"
        self.ws.cell(row=work_exp_header_row, column=3).value = "单位名称"
        self.ws.cell(row=work_exp_header_row, column=4).value = "岗位/职务"
        self.ws.cell(row=work_exp_header_row, column=5).value = "是否邮储银行自主研发工作经验"
        
        # 复制研究生区域的样式（从Row 22开始）
        source_base = 22  # 研究生学历的原始起始行
        for i in range(num_rows):
            self._copy_row_style(source_base + i, insert_row + i)
        
        # 标题行合并 A:F
        self.ws.merge_cells(
            start_row=insert_row, start_column=1,
            end_row=insert_row, end_column=6
        )
        
        # 填充博士学历标签
        self.ws.cell(row=insert_row, column=1).value = "博士学历"
        
        # 填充字段标签
        self.ws.cell(row=insert_row + 1, column=1).value = "入学时间"
        self.ws.cell(row=insert_row + 1, column=3).value = "毕业院校"
        self.ws.cell(row=insert_row + 2, column=1).value = "毕业时间"
        self.ws.cell(row=insert_row + 2, column=3).value = "专业"
        self.ws.cell(row=insert_row + 3, column=1).value = "毕业证编号"
        self.ws.cell(row=insert_row + 4, column=1).value = "毕业证学信网在线验证码"
        self.ws.cell(row=insert_row + 6, column=1).value = "学位证编号"
        self.ws.cell(row=insert_row + 7, column=1).value = "学位证学信网在线验证码"
        
        # 填充博士学历数据
        enrollment_date = self._format_date(edu.get("enrollment_date"))
        university = edu.get("university")
        graduation_date = self._format_date(edu.get("graduation_date"))
        major = edu.get("major")
        diploma_number = edu.get("diploma_number")
        diploma_code = edu.get("diploma_verification_code")
        degree_number = edu.get("degree_number")
        degree_code = edu.get("degree_verification_code")
        
        # 填充数据并处理null值高亮
        cells_data = [
            (insert_row + 1, 2, enrollment_date),
            (insert_row + 1, 4, university),
            (insert_row + 2, 2, graduation_date),
            (insert_row + 2, 4, major),
            (insert_row + 3, 2, diploma_number),
            (insert_row + 4, 2, diploma_code),
            (insert_row + 6, 2, degree_number),
            (insert_row + 7, 2, degree_code),
        ]
        
        for row, col, value in cells_data:
            cell = self.ws.cell(row=row, column=col)
            if value is None or value == "":
                cell.value = "【待补充】"
                cell.fill = self.HIGHLIGHT_FILL
                cell.font = self.HIGHLIGHT_FONT
            else:
                cell.value = value
    
    def fill_work_experience(self, data: dict) -> None:
        """填充工作经历（动态行）"""
        work_exp = data.get("work_experience", [])
        if not work_exp:
            return
        
        # 工作经历表头在第31行，数据从第33行开始
        # 模板中有示例行（33-34），我们直接覆盖并根据需要插入
        data_start_row = self._get_actual_row(33)
        num_items = len(work_exp)
        
        # 模板有2行示例，如果数据更多则插入额外行
        template_rows = 2
        extra_rows = max(0, num_items - template_rows)
        if extra_rows > 0:
            # 在数据区域末尾插入新行
            insert_position = data_start_row + template_rows
            self.ws.insert_rows(insert_position, extra_rows)
            # 复制样式到新插入的行
            for i in range(extra_rows):
                self._copy_row_style(data_start_row, insert_position + i)
            self.row_offset += extra_rows
        
        # 在填充数据前，取消工作经历数据区域的合并单元格（保留表头和标题的合并）
        # 数据区域是从 data_start_row 开始的 num_items 行
        merged_ranges_to_remove = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # 只取消数据行（不包括表头）的 B:F 合并
            if (merged_range.min_row >= data_start_row and 
                merged_range.min_row < data_start_row + num_items and
                merged_range.min_col >= 2 and merged_range.max_col >= 3):
                merged_ranges_to_remove.append(merged_range)
        
        for merged_range in merged_ranges_to_remove:
            try:
                self.ws.unmerge_cells(str(merged_range))
            except:
                pass
        
        # 填充所有工作经历数据
        for i, exp in enumerate(work_exp):
            row = data_start_row + i
            
            start_date = self._format_date(exp.get("start_date"))
            end_date = self._format_date(exp.get("end_date"))
            company = exp.get("company")
            position = exp.get("position")
            is_psbc = exp.get("is_psbc_independent_dev")
            
            self.ws.cell(row=row, column=1).value = start_date if start_date else "【待补充】"
            self.ws.cell(row=row, column=2).value = end_date if end_date else "【待补充】"
            self.ws.cell(row=row, column=3).value = company if company else "【待补充】"
            self.ws.cell(row=row, column=4).value = position if position else "【待补充】"
            
            if is_psbc is None:
                self.ws.cell(row=row, column=5).value = "【待补充】"
            else:
                self.ws.cell(row=row, column=5).value = "是" if is_psbc else "否"
            
            # 对null值应用高亮
            for col in range(1, 6):
                cell = self.ws.cell(row=row, column=col)
                if cell.value == "【待补充】":
                    cell.fill = self.HIGHLIGHT_FILL
                    cell.font = self.HIGHLIGHT_FONT
        
        # 如果工作经历数量少于模板示例行数，清空多余的示例行
        if num_items < template_rows:
            for i in range(num_items, template_rows):
                row = data_start_row + i
                for col in range(1, 6):
                    self.ws.cell(row=row, column=col).value = None

    def fill_project_experience(self, data: dict) -> None:
        """填充项目经历（动态行）"""
        project_exp = data.get("project_experience", [])
        if not project_exp:
            return
        
        # 项目经历原始位置是36行表头，38行开始数据
        # 注意：此时row_offset已经包含了工作经历插入的行数
        data_start_row = self._get_actual_row(38)
        num_items = len(project_exp)
        
        # 模板有2行示例，如果数据更多则插入额外行
        template_rows = 2
        extra_rows = max(0, num_items - template_rows)
        
        if extra_rows > 0:
            insert_position = data_start_row + template_rows
            
            # 在插入行之前，先取消可能受影响的合并单元格
            merged_ranges_to_remove = []
            for merged_range in list(self.ws.merged_cells.ranges):
                # 检查合并单元格是否在插入位置之后
                if merged_range.min_row >= insert_position:
                    merged_ranges_to_remove.append(merged_range)
            
            for merged_range in merged_ranges_to_remove:
                try:
                    self.ws.unmerge_cells(str(merged_range))
                except:
                    pass
            
            # 插入新行
            self.ws.insert_rows(insert_position, extra_rows)
            
            # 复制样式到新插入的行
            for i in range(extra_rows):
                self._copy_row_style(data_start_row, insert_position + i)
            
            self.row_offset += extra_rows
        
        # 填充所有项目经历数据 - 使用直接写入而不是通过 _fill_cell
        for i, exp in enumerate(project_exp):
            row = data_start_row + i
            
            start_date = self._format_date(exp.get("start_date"))
            end_date = self._format_date(exp.get("end_date"))
            project_name = exp.get("project_name")
            description = exp.get("description")
            role = exp.get("role")
            is_psbc = exp.get("is_psbc_independent_dev")
            
            # 直接写入单元格值
            cell1 = self.ws.cell(row=row, column=1)
            cell1.value = start_date if start_date else "【待补充】"
            
            cell2 = self.ws.cell(row=row, column=2)
            cell2.value = end_date if end_date else "【待补充】"
            
            cell3 = self.ws.cell(row=row, column=3)
            cell3.value = project_name if project_name else "【待补充】"
            
            cell4 = self.ws.cell(row=row, column=4)
            cell4.value = description if description else "【待补充】"
            
            cell5 = self.ws.cell(row=row, column=5)
            cell5.value = role if role else "【待补充】"
            
            cell6 = self.ws.cell(row=row, column=6)
            if is_psbc is None:
                cell6.value = "【待补充】"
            else:
                cell6.value = "是" if is_psbc else "否"
            
            # 对null值应用高亮
            for cell in [cell1, cell2, cell3, cell4, cell5, cell6]:
                if cell.value == "【待补充】":
                    cell.fill = self.HIGHLIGHT_FILL
                    cell.font = self.HIGHLIGHT_FONT
    
    def fill_technical_skills(self, data: dict) -> None:
        """填充技术特长"""
        skills = data.get("technical_skills", {})
        
        # 技术特长区域原始位置是41-44行
        base_row = 41
        
        # 掌握语言 - B42
        self._fill_cell(base_row + 1, 2, skills.get("programming_languages"))
        # 掌握技能 - B43
        self._fill_cell(base_row + 2, 2, skills.get("skills"))
        # 专业证书 - B44
        self._fill_cell(base_row + 3, 2, skills.get("certifications"))
    
    def generate(self, resume_data: dict, output_path: str) -> str:
        """
        生成填充后的简历文档
        
        Args:
            resume_data: 解析后的简历JSON数据
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
        """
        print("开始生成简历模板...")
        
        # 按顺序填充各部分（顺序很重要，因为要正确计算行偏移）
        print("  [1/6] 填充基本信息...")
        self.fill_basic_info(resume_data)
        
        print("  [2/6] 填充个人信息...")
        self.fill_personal_info(resume_data)
        
        print("  [3/6] 填充学历信息...")
        self.fill_education(resume_data)
        
        print("  [4/6] 填充工作经历...")
        self.fill_work_experience(resume_data)
        
        print("  [5/6] 填充项目经历...")
        self.fill_project_experience(resume_data)
        
        print("  [6/6] 填充技术特长...")
        self.fill_technical_skills(resume_data)
        
        # 保存文件
        output_path = Path(output_path)
        self.wb.save(output_path)
        print(f"\n✅ 简历模板已生成: {output_path}")
        
        return str(output_path)


def main():
    """主函数"""
    import sys
    
    # 路径配置
    base_dir = Path(__file__).parent
    template_path = base_dir / "Templates" / "template.xlsx"
    
    # 支持命令行参数传入 JSON 路径
    if len(sys.argv) > 1:
        parsed_json_path = sys.argv[1]
        # 如果是相对路径，转换为绝对路径
        if not os.path.isabs(parsed_json_path):
            parsed_json_path = base_dir / parsed_json_path
        else:
            parsed_json_path = Path(parsed_json_path)
    else:
        # 默认路径（用于调试）
        parsed_json_path = base_dir / "Resumes" / "王保利_20250706_191243_parsed.json"
    
    # 输出路径：将 _parsed.json 替换为 _filled.xlsx
    output_path = Path(str(parsed_json_path).replace("_parsed.json", "_filled.xlsx"))
    
    if not parsed_json_path.exists():
        print(f"错误: 文件不存在 - {parsed_json_path}")
        sys.exit(1)
    
    print("=" * 60)
    print("简历模板生成器")
    print("=" * 60)
    
    # 读取解析后的JSON数据
    print(f"\n读取解析数据: {parsed_json_path}")
    with open(parsed_json_path, "r", encoding="utf-8") as f:
        resume_data = json.load(f)
    
    # 检查是否有错误
    if "error" in resume_data:
        print(f"❌ 解析数据包含错误: {resume_data['error']}")
        return
    
    # 生成模板
    print(f"使用模板: {template_path}")
    generator = ResumeTemplateGenerator(str(template_path))
    generator.generate(resume_data, str(output_path))
    
    print("\n" + "=" * 60)
    print("生成完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()
