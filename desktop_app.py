#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Resume Parser - Desktop Application
PyWebView + Flask Implementation
"""

import os
import sys
import json
import tempfile
import threading
import webbrowser
import uuid
import time
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import Optional, List
from concurrent.futures import ThreadPoolExecutor, as_completed
import base64
import zipfile
import io

# Windows 控制台编码修复
if sys.platform == 'win32':
    # 设置标准输出编码为 UTF-8
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    # 设置环境变量
    os.environ['PYTHONIOENCODING'] = 'utf-8'

from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS

# 导入核心模块
from resume_parser import extract_text_from_resume, parse_resume_with_llm
from resume_template_generator import ResumeTemplateGenerator
from license_manager import (
    get_license_manager, 
    validate_license, 
    consume_quota,
    check_startup_license,
    save_local_license,
    get_local_license,
    log_resume_result
)

# ============================================================
# Flask 应用
# ============================================================

# 默认端口（5000 被 macOS AirPlay Receiver 占用）
DEFAULT_PORT = 5050

# PyInstaller 单文件模式资源路径处理
def get_resource_path(relative_path):
    """获取资源文件的绝对路径，兼容 PyInstaller 打包"""
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 打包后的临时目录
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

# 静态文件和模板路径
static_folder = get_resource_path('static')
templates_folder = get_resource_path('Templates')

app = Flask(__name__, static_folder=static_folder, static_url_path='')
CORS(app)

# 全局任务存储
tasks = {}
template_path = Path(templates_folder) / "template.xlsx"

# 并行处理配置
MAX_PARALLEL_WORKERS = 3  # 最大并行数，避免 API 限流


@dataclass
class Task:
    id: str
    filename: str
    status: str = "pending"  # pending, processing, completed, error
    progress: int = 0
    message: str = ""
    result: Optional[dict] = None
    excel_path: Optional[str] = None
    remaining_quota: Optional[str] = None  # 剩余配额，用于前端实时更新
    
    def to_dict(self):
        return asdict(self)


@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/favicon.ico')
def favicon():
    return '', 204  # 返回空响应


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    
    # 检查格式
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ['.pdf', '.docx']:
        return jsonify({'error': f'不支持的格式: {suffix}'}), 400
    
    # 保存临时文件
    temp_dir = tempfile.mkdtemp()
    temp_path = os.path.join(temp_dir, file.filename)
    file.save(temp_path)
    
    # 创建任务 - 使用 UUID 确保唯一性
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    task = Task(id=task_id, filename=file.filename)
    tasks[task_id] = {
        'task': task,
        'temp_path': temp_path
    }
    
    return jsonify({
        'id': task_id,
        'filename': file.filename,
        'status': 'pending'
    })


@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    """获取所有任务"""
    return jsonify([t['task'].to_dict() for t in tasks.values()])


@app.route('/api/tasks/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    """删除任务"""
    if task_id in tasks:
        # 清理临时文件
        temp_path = tasks[task_id].get('temp_path')
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)
        del tasks[task_id]
    return jsonify({'success': True})


@app.route('/api/process', methods=['POST'])
def process_files():
    """开始处理所有待处理的任务（并行处理）"""
    pending_tasks = [t for t in tasks.values() if t['task'].status == 'pending']
    
    if not pending_tasks:
        return jsonify({'error': '没有待处理的任务'}), 400
    
    num_tasks = len(pending_tasks)
    
    # 立即将任务状态设置为 processing，让前端立即看到变化
    for item in pending_tasks:
        item['task'].status = 'processing'
        item['task'].progress = 2
        item['task'].message = '准备中...'
    
    # 在后台线程中进行配额检查和处理
    def process_all_parallel():
        total_start = time.time()
        
        # 先检查并扣减配额
        license_code = get_local_license()
        remaining_quota = None
        
        if license_code:
            quota_result = consume_quota(license_code, num_tasks)
            if not quota_result.get('success'):
                # 配额不足，将所有任务标记为错误
                error_msg = f"配额不足: {quota_result.get('message', '未知错误')}"
                for item in pending_tasks:
                    item['task'].status = 'error'
                    item['task'].message = error_msg
                return
            remaining_quota = quota_result.get('remaining_quota') or str(quota_result.get('remaining', ''))
        
        print(f"\n{'='*60}")
        print(f"[START] Processing {num_tasks} resumes (max workers: {MAX_PARALLEL_WORKERS})")
        print(f"{'='*60}")
        
        # 使用线程池并行处理
        with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
            # 提交所有任务（传入剩余配额信息）
            future_to_task = {
                executor.submit(process_single_task, item, remaining_quota): item 
                for item in pending_tasks
            }
            
            # 等待任务完成并处理结果
            completed_count = 0
            for future in as_completed(future_to_task):
                item = future_to_task[future]
                task = item['task']
                completed_count += 1
                try:
                    future.result()  # 获取结果，捕获异常
                    print(f"[OK] [{completed_count}/{num_tasks}] {task.filename} done")
                except Exception as e:
                    task.status = 'error'
                    task.message = str(e)
                    print(f"[ERR] [{completed_count}/{num_tasks}] {task.filename} failed: {e}")
        
        total_time = time.time() - total_start
        print(f"\n{'='*60}")
        print(f"[DONE] All completed! Total time: {total_time:.1f}s")
        print(f"[STAT] Avg: {total_time/num_tasks:.1f}s per resume")
        print(f"{'='*60}\n")
    
    thread = threading.Thread(target=process_all_parallel)
    thread.start()
    
    return jsonify({'message': f'开始并行处理 {len(pending_tasks)} 个文件 (最大并发: {MAX_PARALLEL_WORKERS})'})


def process_single_task(item, remaining_quota=None):
    """处理单个任务"""
    task = item['task']
    temp_path = item['temp_path']
    timing = {}  # 性能计时
    
    try:
        total_start = time.time()
        
        # 配额已在 process_files 中一次性扣减，这里更新显示
        if remaining_quota:
            task.remaining_quota = remaining_quota
        
        task.progress = 10
        task.message = '正在提取文本...'
        
        # 提取文本
        t0 = time.time()
        text = extract_text_from_resume(temp_path)
        timing['text_extraction'] = time.time() - t0
        print(f"[PERF] Text extraction: {timing['text_extraction']:.2f}s")
        
        task.progress = 30
        task.message = '正在调用 AI 解析...'
        
        # AI 解析
        t0 = time.time()
        result = parse_resume_with_llm(text)
        timing['ai_parsing'] = time.time() - t0
        print(f"[PERF] AI parsing: {timing['ai_parsing']:.2f}s")
        
        task.progress = 70
        task.result = result
        task.message = '正在生成 Excel...'
        
        # 生成 Excel
        t0 = time.time()
        generator = ResumeTemplateGenerator(str(template_path))
        excel_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        generator.generate(result, excel_temp.name)
        timing['excel_generation'] = time.time() - t0
        print(f"[PERF] Excel generation: {timing['excel_generation']:.2f}s")
        
        task.excel_path = excel_temp.name
        
        timing['total'] = time.time() - total_start
        print(f"[PERF] Total: {timing['total']:.2f}s")
        print(f"[PERF] Distribution: Text {timing['text_extraction']/timing['total']*100:.1f}% | AI {timing['ai_parsing']/timing['total']*100:.1f}% | Excel {timing['excel_generation']/timing['total']*100:.1f}%")
        
        task.progress = 100
        task.status = 'completed'
        task.message = f"处理完成 (耗时{timing['total']:.1f}s)"
        
        # 后台异步上传解析结果到云端
        license_code = get_local_license()
        if license_code and result:
            log_resume_result(
                license_code=license_code,
                filename=task.filename,
                result_json=result,
                status="success"
            )
        
    except Exception as e:
        task.status = 'error'
        task.message = str(e)
        
        # 记录错误日志
        license_code = get_local_license()
        if license_code:
            log_resume_result(
                license_code=license_code,
                filename=task.filename,
                result_json={},
                status="error",
                error_message=str(e)
            )


@app.route('/api/tasks/<task_id>/status', methods=['GET'])
def get_task_status(task_id):
    """获取任务状态"""
    if task_id not in tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    return jsonify(tasks[task_id]['task'].to_dict())


@app.route('/api/tasks/<task_id>/download', methods=['GET'])
def download_excel(task_id):
    """下载 Excel 文件"""
    if task_id not in tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    task = tasks[task_id]['task']
    if not task.excel_path or not os.path.exists(task.excel_path):
        return jsonify({'error': '文件不存在'}), 404
    
    # 获取姓名作为文件名
    name = "简历"
    if task.result:
        name = task.result.get("基本信息", {}).get("姓名", Path(task.filename).stem)
    
    return send_file(
        task.excel_path,
        as_attachment=True,
        download_name=f"{name}_简历.xlsx"
    )


@app.route('/api/download-all', methods=['GET'])
def download_all():
    """批量下载所有完成的任务"""
    completed = [t for t in tasks.values() 
                 if t['task'].status == 'completed' and t['task'].excel_path]
    
    if not completed:
        return jsonify({'error': '没有可下载的文件'}), 400
    
    # 创建 ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in completed:
            task = item['task']
            if os.path.exists(task.excel_path):
                name = "简历"
                if task.result:
                    name = task.result.get("基本信息", {}).get("姓名", Path(task.filename).stem)
                zf.write(task.excel_path, f"{name}_简历.xlsx")
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"简历批量导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    )


@app.route('/api/tasks/<task_id>/result', methods=['GET'])
def get_result(task_id):
    """获取解析结果 JSON"""
    if task_id not in tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    task = tasks[task_id]['task']
    if not task.result:
        return jsonify({'error': '暂无结果'}), 404
    
    return jsonify(task.result)


@app.route('/api/tasks/<task_id>/excel-preview', methods=['GET'])
def get_excel_preview(task_id):
    """获取 Excel 预览数据"""
    from openpyxl import load_workbook
    
    if task_id not in tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    task = tasks[task_id]['task']
    if not task.excel_path or not os.path.exists(task.excel_path):
        return jsonify({'error': '文件不存在'}), 404
    
    try:
        wb = load_workbook(task.excel_path)
        ws = wb.active
        
        # 读取所有数据
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):  # 最多100行
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                elif hasattr(value, 'strftime'):  # 日期类型
                    value = value.strftime('%Y-%m-%d')
                else:
                    value = str(value)
                row_data.append(value)
            rows.append(row_data)
        
        # 获取合并单元格信息
        merged_cells = []
        for merged_range in ws.merged_cells.ranges:
            merged_cells.append({
                'startRow': merged_range.min_row - 1,
                'endRow': merged_range.max_row - 1,
                'startCol': merged_range.min_col - 1,
                'endCol': merged_range.max_col - 1
            })
        
        return jsonify({
            'rows': rows,
            'mergedCells': merged_cells,
            'maxCol': ws.max_column
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# 授权管理 API
# ============================================================

@app.route('/api/license/status', methods=['GET'])
def get_license_status():
    """获取当前授权状态"""
    result = check_startup_license()
    return jsonify(result)


@app.route('/api/license/validate', methods=['POST'])
def validate_license_api():
    """验证授权码"""
    data = request.get_json()
    code = data.get('code', '')
    
    result = validate_license(code)
    
    # 如果验证成功，保存到本地
    if result.get('valid'):
        save_local_license(code)
    
    return jsonify(result)


# ============================================================
# 启动函数
# ============================================================

def run_flask():
    """运行 Flask 服务"""
    try:
        app.run(host='127.0.0.1', port=DEFAULT_PORT, debug=False, threaded=True)
    except Exception as e:
        print(f"Flask startup failed: {e}")


def wait_for_server(port, timeout=30):
    """等待服务器启动"""
    import socket
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            if result == 0:
                return True
        except:
            pass
        time.sleep(0.5)
    return False


def run_webview():
    """Run PyWebView window"""
    import webview
    
    # Start Flask in background thread
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    # Wait for Flask to start (max 30 seconds)
    print(f"Waiting for server at http://127.0.0.1:{DEFAULT_PORT} ...")
    if not wait_for_server(DEFAULT_PORT, timeout=30):
        print("Error: Flask server startup timeout!")
        return
    print("Server started!")
    
    # Create window
    webview.create_window(
        'CYBER RESUME PARSER v2.0',
        f'http://127.0.0.1:{DEFAULT_PORT}',
        width=1000,
        height=750,
        min_size=(800, 600),
        resizable=True,
        background_color='#0f0f1a'
    )
    
    # Windows uses edgechromium, macOS uses default cocoa
    import platform
    if platform.system() == 'Windows':
        webview.start(gui='edgechromium')
    else:
        webview.start()


def run_browser():
    """Run in browser (dev mode)"""
    print("=" * 50)
    print("AI Resume Parser - Dev Mode")
    print("=" * 50)
    print(f"Open browser: http://127.0.0.1:{DEFAULT_PORT}")
    print("Press Ctrl+C to stop")
    print("=" * 50)
    
    # 只在主进程打开浏览器（debug 模式会重启子进程）
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        threading.Timer(1.5, lambda: webbrowser.open(f'http://127.0.0.1:{DEFAULT_PORT}')).start()
    
    app.run(host='127.0.0.1', port=DEFAULT_PORT, debug=True)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', choices=['browser', 'desktop'], default='desktop',
                        help='运行模式: browser(浏览器) 或 desktop(桌面窗口)')
    args, _ = parser.parse_known_args()  # 忽略未知参数
    
    if args.mode == 'desktop':
        try:
            import webview
            run_webview()
        except ImportError:
            print("请先安装 pywebview: pip install pywebview")
            sys.exit(1)
    else:
        run_browser()
